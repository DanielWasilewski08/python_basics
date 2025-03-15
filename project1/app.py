# you have to download this file to see created chart
# if you want to see corrected price, reload this page
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook("project1/"+filename)
    sheet = wb["Sheet1"]
#creating corrected price with 90% discount
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

#establishing sizes
    values = Reference(sheet, 
              min_row=2, 
              max_row=sheet.max_row,
              min_col=4,
              max_col=7)
#creating chart
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "e2")
#writing info about downloading excel file in e1 cell
    chart_info = "you have to download excel file to see chart"
    chart_info_cell = sheet.cell(1,5)
    chart_info_cell.value = chart_info
#saving file    
    wb.save("project1/"+filename)
#starting function
process_workbook("transactions.xlsx")